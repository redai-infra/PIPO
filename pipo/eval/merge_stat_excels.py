from __future__ import annotations

import argparse
import re
from copy import copy
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


DATA_START_COL = 5
METRIC_PER_BUCKET = 3  # pass@1, pass@k, finished_num


# ── path classification ──────────────────────────────────────────────────────

_OURS_RE = re.compile(
    r"outputs/[^/]+/[^/]+/checkpoint-[^/]+/eval/[^/]+/stats\.xlsx$"
)
_BASELINE_RE = re.compile(
    r"outputs/[^/]+/[^/]+/[^/]+/stats\.xlsx$"
)


def classify(path: Path) -> str:
    """Return 'ours', 'baseline', or 'other' for a given stats.xlsx path."""
    s = str(path).replace("\\", "/")
    if _OURS_RE.search(s):
        return "ours"
    if _BASELINE_RE.search(s):
        return "baseline"
    return "other"


def discover_inputs(root: Path) -> List[Path]:
    """Glob stats.xlsx under *root* and return them in [baselines..., ours...] order."""
    found = list(root.rglob("stats.xlsx"))
    baselines, ours, others = [], [], []
    for p in found:
        kind = classify(p)
        if kind == "baseline":
            baselines.append(p)
        elif kind == "ours":
            ours.append(p)
        else:
            others.append(p)
    baselines.sort()
    ours.sort()
    others.sort()
    if others:
        print(f"[warn] {len(others)} stats.xlsx don't match either pattern, skipped:")
        for p in others:
            print(f"  - {p}")
    print(f"[discover] baseline={len(baselines)}, ours={len(ours)} (under {root})")
    return baselines + ours


# ── helpers ──────────────────────────────────────────────────────────────────


def _copy_cell(src_ws: Worksheet, dst_ws: Worksheet,
               src_row: int, src_col: int,
               dst_row: int, dst_col: int) -> None:
    src = src_ws.cell(row=src_row, column=src_col)
    dst = dst_ws.cell(row=dst_row, column=dst_col)
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)


def _copy_merged_cells(src_ws: Worksheet, dst_ws: Worksheet) -> None:
    for mc_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(
            start_row=mc_range.min_row, start_column=mc_range.min_col,
            end_row=mc_range.max_row, end_column=mc_range.max_col,
        )


def _copy_column_widths(src_ws: Worksheet, dst_ws: Worksheet) -> None:
    for col_letter, col_dim in src_ws.column_dimensions.items():
        if col_dim.width is not None:
            dst_ws.column_dimensions[col_letter].width = col_dim.width


# ── core merge logic ─────────────────────────────────────────────────────────


def _append_data_row(
    src_ws: Worksheet, dst_ws: Worksheet, target_row: int, is_ours: bool,
) -> None:
    """Copy row 3 (the single data row) from *src_ws* to *dst_ws*.

    When ``is_ours`` is True, the length-bucket triples (cols ``DATA_START_COL``
    onward) are shifted left by one bucket (= ``METRIC_PER_BUCKET`` cols),
    because ours-method runs at ~2x token compression: their 4K performance
    is compared against baseline 2K, 8K vs 4K, ..., 128K vs 64K. The 128K
    destination column is left empty for ours rows.
    """
    max_col = src_ws.max_column

    # 1) Meta cols (training config / inference config / avg_token / avg_pad) -- as-is.
    for col in range(1, min(DATA_START_COL, max_col + 1)):
        _copy_cell(src_ws, dst_ws, 3, col, target_row, col)

    # 2) Length-bucket cols.
    if is_ours:
        for src_col in range(DATA_START_COL + METRIC_PER_BUCKET, max_col + 1):
            dst_col = src_col - METRIC_PER_BUCKET
            _copy_cell(src_ws, dst_ws, 3, src_col, target_row, dst_col)
        # The last bucket (128K) is intentionally left blank for ours rows.
    else:
        for col in range(DATA_START_COL, max_col + 1):
            _copy_cell(src_ws, dst_ws, 3, col, target_row, col)


def merge_stat_excels(input_paths: List[Path], output_path: Path) -> None:
    if not input_paths:
        raise ValueError("no input files")

    # 1) Discover sheet names from the first file
    first_wb = load_workbook(input_paths[0], data_only=True)
    sheet_names = first_wb.sheetnames
    first_wb.close()

    # 2) Output workbook
    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    output_sheets: dict[str, Worksheet] = {
        name: out_wb.create_sheet(title=name) for name in sheet_names
    }
    next_row_per_sheet: dict[str, int] = {name: 3 for name in sheet_names}
    headers_copied: set[str] = set()

    # 3) Copy headers (once) + append data row (per input, per sheet)
    for idx, inp_path in enumerate(input_paths):
        is_ours = classify(inp_path) == "ours"
        tag = "ours " if is_ours else "base "
        print(f"  [{idx + 1}/{len(input_paths)}] {tag}{inp_path}")
        wb = load_workbook(inp_path, data_only=True)

        for name in sheet_names:
            if name not in wb.sheetnames:
                print(f"    [warn] sheet '{name}' missing, skipped")
                continue
            src_ws = wb[name]
            dst_ws = output_sheets[name]

            if name not in headers_copied:
                for row in (1, 2):
                    for col in range(1, src_ws.max_column + 1):
                        _copy_cell(src_ws, dst_ws, row, col, row, col)
                _copy_merged_cells(src_ws, dst_ws)
                _copy_column_widths(src_ws, dst_ws)
                headers_copied.add(name)

            target_row = next_row_per_sheet[name]
            _append_data_row(src_ws, dst_ws, target_row, is_ours)
            next_row_per_sheet[name] += 1

        wb.close()

    # 4) Freeze headers
    for ws in output_sheets.values():
        ws.freeze_panes = "A9"

    # 5) Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)
    print(f"[done] merged {len(input_paths)} files -> {output_path}")
    print(f"  sheets: {sheet_names}")


# ── CLI ──────────────────────────────────────────────────────────────────────


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Merge stats.xlsx files (baseline rows on top, ours below).",
    )
    p.add_argument(
        "inputs", nargs="*", type=Path,
        help="Explicit stats.xlsx paths (order preserved). If omitted, auto-glob from --root.",
    )
    p.add_argument(
        "--root", type=Path, default=Path("outputs"),
        help="Root dir to glob when no explicit inputs are given (default: outputs).",
    )
    p.add_argument(
        "-o", "--output", type=Path, default=None,
        help="Output xlsx path (default: <root>/all_stats_summary.xlsx).",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()

    if args.inputs:
        input_paths = args.inputs
        for p in input_paths:
            if not p.is_file():
                print(f"[error] not found: {p}")
                return
    else:
        if not args.root.is_dir():
            print(f"[error] root not a directory: {args.root}")
            return
        input_paths = discover_inputs(args.root)
        if not input_paths:
            print(f"[error] no stats.xlsx found under {args.root}")
            return

    output = args.output if args.output else args.root / "all_stats_summary.xlsx"
    merge_stat_excels(input_paths, output)


if __name__ == "__main__":
    main()
