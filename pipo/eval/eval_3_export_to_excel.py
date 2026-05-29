import argparse
import os
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

sys.path.append(os.getcwd())

from pipo.eval.eval_utils import (
    EXCEL_BENCHMARKS,
    LENGTH_ORDER,
    compute_statistics,
    load_jsonl_dedup,
    safe_mean,
    write_statistics_json,
)


# ---- path parsing ----

_OURS_RE = re.compile(
    r".*/outputs/(?P<model>[^/]+)/(?P<method>[^/]+)/"
    r"checkpoint-(?P<step>[^-]+)(?:-merged)?/eval/(?P<infer_cfg>[^/]+)/?$"
)
_BASELINE_RE = re.compile(
    r".*/outputs/(?P<model>[^/]+)/(?P<method>[^/]+)/(?P<infer_cfg>[^/]+)/?$"
)


def parse_exp_dir(exp_dir: Path) -> dict:
    s = str(exp_dir.resolve()).rstrip("/")
    m = _OURS_RE.match(s)
    if m:
        d = m.groupdict()
        return {
            "model": d["model"], "method": d["method"],
            "step": d["step"], "infer_cfg": d["infer_cfg"],
            "label_train": f"{d['model']}/{d['method']} (ckpt-{d['step']})",
            "label_infer": d["infer_cfg"],
        }
    m = _BASELINE_RE.match(s)
    if m:
        d = m.groupdict()
        return {
            "model": d["model"], "method": d["method"],
            "step": "", "infer_cfg": d["infer_cfg"],
            "label_train": f"{d['model']}/{d['method']}",
            "label_infer": d["infer_cfg"],
        }
    raise ValueError(f"Cannot parse exp_dir from {exp_dir}")


# ---- aggregate ----

def compute_overall(per_bench: dict[str, dict]) -> dict:
    """Average pass@1 / pass@k / finished_num per length bucket across benchmarks."""
    out: dict[str, dict] = {}
    for length in LENGTH_ORDER:
        bucket = {"pass@1": [], "pass@k": [], "finished_num": []}
        for bench in EXCEL_BENCHMARKS:
            stats = per_bench.get(bench)
            if stats is None:
                continue
            ml = stats["multi_length_stats"].get(length)
            if ml is None:
                continue
            for key in bucket:
                bucket[key].append(float(ml[key]))
        out[length] = {
            "pass@1": safe_mean(bucket["pass@1"]),
            "pass@k": safe_mean(bucket["pass@k"]),
            "finished_num": safe_mean(bucket["finished_num"]),
        }
    avg_token = safe_mean([
        per_bench[b]["avg_token_length-all"] for b in EXCEL_BENCHMARKS if b in per_bench
    ])
    avg_pad = safe_mean([
        per_bench[b]["avg_pad_length-all"] for b in EXCEL_BENCHMARKS if b in per_bench
    ])
    return {
        "question_num": sum(per_bench[b]["question_num"] for b in EXCEL_BENCHMARKS if b in per_bench),
        "avg_token_length-all": avg_token,
        "avg_pad_length-all": avg_pad,
        "multi_length_stats": out,
    }


# ---- excel ----

HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid")
TRAIN_FONT = Font(bold=True)
INFER_FONT = Font(italic=True, color="FF555555")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

METRIC_COLUMNS = ["pass@1", "pass@k", "finished_num"]


def _fmt_acc(v: Any) -> Any:
    if v is None or v == "":
        return ""
    return round(float(v) * 100, 2)


def _fmt_int(v: Any) -> Any:
    if v is None or v == "":
        return ""
    return int(v)


def _write_sheet(ws: Worksheet, exp: dict, stats: dict | None) -> None:
    data_start_col = 5  # cols: train, infer, avg_token, avg_pad, then per-length

    # row 1 / 2 headers
    for col, title in ((1, "training config"), (2, "inference config")):
        c = ws.cell(row=1, column=col, value=title)
        c.font = HEADER_FONT; c.alignment = CENTER; c.fill = HEADER_FILL
        ws.cell(row=2, column=col).fill = SUBHEADER_FILL
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    for col, title in ((3, "avg_token"), (4, "avg_pad")):
        c = ws.cell(row=1, column=col, value=title)
        c.font = HEADER_FONT; c.alignment = CENTER; c.fill = HEADER_FILL
        ws.cell(row=2, column=col).fill = SUBHEADER_FILL
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    for i, length in enumerate(LENGTH_ORDER):
        start_col = data_start_col + i * len(METRIC_COLUMNS)
        end_col = start_col + len(METRIC_COLUMNS) - 1
        c = ws.cell(row=1, column=start_col, value=length)
        c.font = HEADER_FONT; c.alignment = CENTER
        for cc in range(start_col, end_col + 1):
            ws.cell(row=1, column=cc).fill = HEADER_FILL
        if end_col > start_col:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col,
            )
        for j, metric in enumerate(METRIC_COLUMNS):
            col = start_col + j
            cc = ws.cell(row=2, column=col, value=metric)
            cc.font = HEADER_FONT; cc.alignment = CENTER; cc.fill = SUBHEADER_FILL

    # data row
    r = 3
    c1 = ws.cell(row=r, column=1, value=exp["label_train"])
    c1.font = TRAIN_FONT; c1.alignment = LEFT
    c2 = ws.cell(row=r, column=2, value=exp["label_infer"])
    c2.font = INFER_FONT; c2.alignment = LEFT

    if stats is None:
        ws.cell(row=r, column=3, value="N/A").alignment = CENTER
        return

    ws.cell(row=r, column=3, value=_fmt_int(stats["avg_token_length-all"])).alignment = CENTER
    ws.cell(row=r, column=4, value=_fmt_int(stats["avg_pad_length-all"])).alignment = CENTER

    multi = stats["multi_length_stats"]
    for i, length in enumerate(LENGTH_ORDER):
        ml = multi.get(length, {})
        base = data_start_col + i * len(METRIC_COLUMNS)
        ws.cell(row=r, column=base + 0, value=_fmt_acc(ml.get("pass@1"))).alignment = CENTER
        ws.cell(row=r, column=base + 1, value=_fmt_acc(ml.get("pass@k"))).alignment = CENTER
        ws.cell(row=r, column=base + 2, value=_fmt_int(ml.get("finished_num"))).alignment = CENTER

    # column widths
    ws.column_dimensions["A"].width = max(28, len(exp["label_train"]) + 2)
    ws.column_dimensions["B"].width = max(28, len(exp["label_infer"]) + 2)
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    total_cols = (data_start_col - 1) + len(LENGTH_ORDER) * len(METRIC_COLUMNS)
    for c in range(data_start_col, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = f"{get_column_letter(data_start_col)}3"


def _safe_sheet_name(name: str) -> str:
    bad = set(r":\/?*[]")
    cleaned = "".join("_" if ch in bad else ch for ch in name)
    return cleaned[:31] or "sheet"


def export_excel(exp_dir: Path, exp: dict, per_bench: dict[str, dict]) -> Path:
    overall = compute_overall(per_bench)

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title="overall")
    _write_sheet(ws, exp, overall)

    for bench in EXCEL_BENCHMARKS:
        ws = wb.create_sheet(title=_safe_sheet_name(bench))
        _write_sheet(ws, exp, per_bench.get(bench))

    out = exp_dir / "stats.xlsx"
    wb.save(out)
    return out


# ---- main ----

def main() -> None:
    parser = argparse.ArgumentParser(description="Build stats.json + stats.xlsx")
    parser.add_argument("exp_dir_name", type=str)
    args = parser.parse_args()

    exp_dir = Path(args.exp_dir_name)
    if not exp_dir.is_dir():
        print(f"[eval_4] error: {exp_dir} is not a directory")
        sys.exit(1)

    exp = parse_exp_dir(exp_dir)

    per_bench: dict[str, dict] = {}
    for bench in EXCEL_BENCHMARKS:
        jsonl = exp_dir / f"{bench}-results.jsonl"
        if not jsonl.exists():
            print(f"[eval_4] skip {bench}: {jsonl} not found")
            continue
        records = load_jsonl_dedup(jsonl)
        if not records:
            print(f"[eval_4] skip {bench}: empty jsonl")
            continue
        missing = [r["src_item"]["micro_index"] for r in records if r["accuracies"] is None]
        if missing:
            print(
                f"[eval_4] skip {bench}: {len(missing)} records have accuracies=None "
                f"(run stage 1-3 first)"
            )
            continue
        stats = compute_statistics(records)
        out_json = write_statistics_json(exp_dir, bench, stats)
        per_bench[bench] = stats
        ml128 = stats["multi_length_stats"].get("128K", {})
        print(
            f"[eval_4] {bench}: q={stats['question_num']} "
            f"avg_tok={stats['avg_token_length-all']:.0f} "
            f"pass@1[128K]={ml128.get('pass@1', 0):.4f} "
            f"pass@k[128K]={ml128.get('pass@k', 0):.4f} -> {out_json}"
        )

    if not per_bench:
        print(f"[eval_4] no benchmarks available; skip excel")
        return

    xlsx = export_excel(exp_dir, exp, per_bench)
    print(f"[eval_4] wrote {xlsx}")


if __name__ == "__main__":
    main()
